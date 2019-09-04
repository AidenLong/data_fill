# -*- coding utf-8 -*- #

import datetime
from openpyxl import load_workbook


def data_handle():
    # 打开文件，获取excel文件的workbook（工作簿）对象
    file_name = "./data/sa5AUG19_final_bycodeRB.xlsx"
    wb = load_workbook(filename=file_name)  # 打开excel文件

    # 通过sheet索引获得sheet对象
    worksheet = wb['原始表']

    nrows = worksheet.max_row  # 获取该表总行数

    date_2006 = datetime.datetime(2006, 1, 1)
    date_2008 = datetime.datetime(2008, 3, 1)
    date_2013 = datetime.datetime(2013, 3, 1)
    date_2019 = datetime.datetime(2019, 7, 1)
    date_2001 = datetime.datetime(2001, 7, 1)
    date_2004 = datetime.datetime(2004, 7, 1)
    date_2007 = datetime.datetime(2007, 1, 1)
    date_2010 = datetime.datetime(2010, 7, 1)
    date_2017 = datetime.datetime(2017, 1, 1)
    for nrow in range(1, nrows):
        value = worksheet.cell(nrow + 1, 12).value
        if not value or value == '#NAME?':
            city = worksheet.cell(nrow + 1, 23).value
            date = worksheet.cell(nrow + 1, 40).value
            if len(date) == 5 and date.endswith('0'):
                date = date[0:-1] + '1'
            date = datetime.datetime.strptime(str(date), '%Y%m')
            if city == '北京':
                if date <= date_2006:
                    worksheet['L' + str(nrow + 1)] = '国二'
                elif date_2006 < date <= date_2008:
                    # print('国三')
                    worksheet['L' + str(nrow + 1)] = '国三'
                elif date_2008 < date <= date_2013:
                    # print('国四')
                    worksheet['L' + str(nrow + 1)] = '国四'
                elif date_2013 < date <= date_2019:
                    # print('国五')
                    worksheet['L' + str(nrow + 1)] = '国五'
            else:
                if date_2001 < date <= date_2004:
                    # print('国一')
                    worksheet['L' + str(nrow + 1)] = '国一'
                elif date_2004 < date <= date_2007:
                    # print('国二')
                    worksheet['L' + str(nrow + 1)] = '国二'
                elif date_2007 < date <= date_2010:
                    # print('国三')
                    worksheet['L' + str(nrow + 1)] = '国三'
                elif date_2010 < date <= date_2017:
                    # print('国四')
                    worksheet['L' + str(nrow + 1)] = '国四'
                elif date_2017 < date <= date_2019:
                    # print('国五')
                    worksheet['L' + str(nrow + 1)] = '国五'
                elif date_2019 < date:
                    # print('国六')
                    worksheet['L' + str(nrow + 1)] = '国六'

        else:
            value = value.strip()
            if ('三' in value or 'Ⅲ' in value or 'III' in value) and 'OBD' in value:
                # print('国三+OBD')
                worksheet['L' + str(nrow + 1)] = '国三+OBD'
            elif '三' in value or 'Ⅲ' in value or 'III' in value:
                # print('国三')
                worksheet['L' + str(nrow + 1)] = '国三'
            elif '二' in value or 'Ⅱ' in value or 'II' in value:
                # print('国二')
                worksheet['L' + str(nrow + 1)] = '国二'
            elif '四' in value or 'IV' in value or 'Ⅳ' in value:
                # print('国四')
                worksheet['L' + str(nrow + 1)] = '国四'
            elif '五' in value or 'Ⅴ' in value or 'V' in value:
                # print('国五')
                worksheet['L' + str(nrow + 1)] = '国五'
            elif '一' in value:
                # print('国一')
                worksheet['L' + str(nrow + 1)] = '国一'
        print(nrow)
    wb.save('./data/data.xlsx')


if __name__ == '__main__':
    data_handle()
